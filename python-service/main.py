"""
الصرح للتطوير العقاري - Python Analytics Service
FastAPI microservice for analytics and reporting
"""
from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import List, Optional, Any, Dict
from datetime import datetime
import statistics
import io

# openpyxl for rich Excel export
try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

app = FastAPI(
    title="الصرح Analytics API",
    description="خدمة تحليلات بايثون للصرح للتطوير العقاري",
    version="1.0.0"
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ── Models ───────────────────────────────────────────────────────────────────

class UnitData(BaseModel):
    id: str
    price: float
    area: float
    rooms: int
    floor: Optional[str] = None
    type: str
    status: str

class PriceAnalysisRequest(BaseModel):
    units: List[UnitData]

class LeadData(BaseModel):
    date: str
    source: Optional[str] = "website"

class LeadsAnalyticsRequest(BaseModel):
    leads: List[LeadData]

# ── Routes ───────────────────────────────────────────────────────────────────

@app.get("/health")
def health():
    return {"status": "ok", "service": "الصرح Analytics", "timestamp": datetime.now().isoformat()}

@app.post("/analytics/units/price-analysis")
def price_analysis(req: PriceAnalysisRequest):
    """تحليل أسعار الوحدات"""
    units = req.units
    if not units:
        raise HTTPException(status_code=400, detail="لا توجد وحدات للتحليل")

    prices = [u.price for u in units if u.price > 0]
    areas = [u.area for u in units if u.area > 0]
    price_per_sqm = [u.price / u.area for u in units if u.price > 0 and u.area > 0]

    # Group by type
    by_type = {}
    for u in units:
        if u.type not in by_type:
            by_type[u.type] = []
        by_type[u.type].append(u.price)

    type_stats = {
        t: {
            "count": len(prices_list),
            "avg": round(statistics.mean(prices_list)) if prices_list else 0,
            "min": min(prices_list) if prices_list else 0,
            "max": max(prices_list) if prices_list else 0,
        }
        for t, prices_list in by_type.items()
        if prices_list
    }

    # Status distribution
    status_dist = {}
    for u in units:
        status_dist[u.status] = status_dist.get(u.status, 0) + 1

    return {
        "total_units": len(units),
        "price_stats": {
            "min": min(prices) if prices else 0,
            "max": max(prices) if prices else 0,
            "avg": round(statistics.mean(prices)) if prices else 0,
            "median": round(statistics.median(prices)) if prices else 0,
        },
        "area_stats": {
            "min": min(areas) if areas else 0,
            "max": max(areas) if areas else 0,
            "avg": round(statistics.mean(areas), 1) if areas else 0,
        },
        "price_per_sqm": {
            "avg": round(statistics.mean(price_per_sqm)) if price_per_sqm else 0,
            "min": round(min(price_per_sqm)) if price_per_sqm else 0,
            "max": round(max(price_per_sqm)) if price_per_sqm else 0,
        },
        "by_type": type_stats,
        "status_distribution": status_dist,
    }

@app.post("/analytics/leads/trends")
def leads_trends(req: LeadsAnalyticsRequest):
    """تحليل توجهات العملاء المحتملين"""
    leads = req.leads
    if not leads:
        return {"monthly": {}, "total": 0}

    monthly = {}
    source_dist = {}

    for lead in leads:
        try:
            dt = datetime.fromisoformat(lead.date.replace("Z", "+00:00"))
            key = f"{dt.year}-{str(dt.month).zfill(2)}"
            monthly[key] = monthly.get(key, 0) + 1
        except:
            pass

        src = lead.source or "website"
        source_dist[src] = source_dist.get(src, 0) + 1

    return {
        "total": len(leads),
        "monthly": dict(sorted(monthly.items())),
        "source_distribution": source_dist,
        "peak_month": max(monthly, key=monthly.get) if monthly else None,
    }

@app.get("/analytics/price-estimate")
def price_estimate(area: float, rooms: int, unit_type: str = "apartment", floor: int = 1):
    """تقدير سعر وحدة بناءً على المواصفات (بسيط)"""
    # Base prices per sqm by type (EGP)
    base_prices = {
        "apartment": 15000,
        "villa": 25000,
        "studio": 18000,
        "duplex": 20000,
        "penthouse": 30000,
        "office": 22000,
        "shop": 28000,
        "chalet": 12000,
    }

    base_per_sqm = base_prices.get(unit_type, 15000)

    # Adjustments
    room_bonus = (rooms - 2) * 0.05  # +5% per room above 2
    floor_bonus = min(floor * 0.02, 0.15)  # +2% per floor, max 15%

    multiplier = 1 + room_bonus + floor_bonus
    estimated = area * base_per_sqm * multiplier

    return {
        "estimated_price": round(estimated),
        "price_per_sqm": round(base_per_sqm * multiplier),
        "area": area,
        "rooms": rooms,
        "type": unit_type,
        "note": "تقدير تقريبي بناءً على متوسطات السوق",
    }

# ── Excel Export ─────────────────────────────────────────────────────────────

class ColumnDef(BaseModel):
    key: str
    label: str
    type: str = "text"       # text | number | currency | date | percentage | formula
    formula: Optional[str] = None

class RowData(BaseModel):
    cells: Dict[str, Any] = {}

class ExcelExportRequest(BaseModel):
    sheet_name: str = "البيانات"
    ledger_name: Optional[str] = None
    columns: List[ColumnDef]
    rows: List[RowData]
    include_totals: bool = True

def _eval_formula(formula: str, cells: Dict[str, Any]) -> float:
    """Safe formula evaluator matching frontend logic."""
    if not formula:
        return 0.0
    expr = str(formula)
    # replace longest keys first
    for key in sorted(cells.keys(), key=len, reverse=True):
        try:
            num = float(cells[key])
        except (ValueError, TypeError):
            num = 0.0
        expr = expr.replace(key, f"({num})")
    # remove any remaining identifiers
    import re
    expr = re.sub(r"[a-zA-Z_][a-zA-Z0-9_]*", "0", expr)
    if not re.match(r"^[\d\s\+\-\*\/\.\(\)]+$", expr):
        return 0.0
    try:
        result = eval(expr, {"__builtins__": {}})
        return float(result) if result is not None else 0.0
    except Exception:
        return 0.0


@app.post("/export/excel")
def export_excel(req: ExcelExportRequest):
    """
    تصدير بيانات الجدول إلى Excel مع تنسيق احترافي.
    POST /export/excel → returns .xlsx file
    """
    if not OPENPYXL_AVAILABLE:
        raise HTTPException(status_code=501, detail="openpyxl غير متوفر")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = req.sheet_name[:31]  # Excel sheet name max 31 chars
    ws.sheet_view.rightToLeft = True  # RTL direction

    # ── Colours ──────────────────────────────────────────────────────────────
    HEADER_FILL  = PatternFill("solid", fgColor="2D5D89")
    TOTAL_FILL   = PatternFill("solid", fgColor="DBEAFE")
    ALT_ROW_FILL = PatternFill("solid", fgColor="F8FAFC")
    WHITE_FILL   = PatternFill("solid", fgColor="FFFFFF")

    HEADER_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=12)
    TOTAL_FONT   = Font(bold=True, color="2D5D89", name="Calibri", size=11)
    DATA_FONT    = Font(name="Calibri", size=11)
    TITLE_FONT   = Font(bold=True, color="2D5D89", name="Calibri", size=14)

    thin_side = Side(style="thin", color="E5E7EB")
    BORDER = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    HEADER_BORDER = Border(
        left=Side(style="medium", color="1F4566"),
        right=Side(style="medium", color="1F4566"),
        top=Side(style="medium", color="1F4566"),
        bottom=Side(style="medium", color="1F4566"),
    )

    current_row = 1

    # ── Title row ─────────────────────────────────────────────────────────────
    if req.ledger_name:
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=len(req.columns))
        cell = ws.cell(row=current_row, column=1, value=req.ledger_name)
        cell.font = TITLE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[current_row].height = 30
        current_row += 1

    # Sheet name sub-title
    ws.merge_cells(start_row=current_row, start_column=1,
                   end_row=current_row, end_column=len(req.columns))
    cell = ws.cell(row=current_row, column=1, value=req.sheet_name)
    cell.font = Font(bold=True, color="334155", name="Calibri", size=12)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[current_row].height = 22
    current_row += 1

    # Date
    ws.merge_cells(start_row=current_row, start_column=1,
                   end_row=current_row, end_column=len(req.columns))
    cell = ws.cell(row=current_row, column=1,
                   value=f"تاريخ التصدير: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    cell.font = Font(color="64748B", name="Calibri", size=10)
    cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[current_row].height = 18
    current_row += 1
    current_row += 1  # blank row

    header_row = current_row

    # ── Header row ────────────────────────────────────────────────────────────
    for ci, col in enumerate(req.columns, start=1):
        cell = ws.cell(row=header_row, column=ci, value=col.label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = HEADER_BORDER
    ws.row_dimensions[header_row].height = 28
    current_row += 1

    # ── Data rows ─────────────────────────────────────────────────────────────
    totals: Dict[str, float] = {col.key: 0.0 for col in req.columns}

    for ri, row in enumerate(req.rows):
        fill = ALT_ROW_FILL if ri % 2 == 0 else WHITE_FILL
        for ci, col in enumerate(req.columns, start=1):
            raw_val = row.cells.get(col.key, "")

            if col.type == "formula":
                val = _eval_formula(col.formula or "", row.cells)
            elif col.type in ("number", "currency", "percentage"):
                try:
                    val = float(raw_val) if raw_val != "" else ""
                except (ValueError, TypeError):
                    val = raw_val
            else:
                val = raw_val

            cell = ws.cell(row=current_row, column=ci, value=val)
            cell.font = DATA_FONT
            cell.fill = fill
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="right" if col.type == "text" else "center",
                                       vertical="center")

            # Number formats
            if col.type == "currency" and isinstance(val, (int, float)):
                cell.number_format = '#,##0 "ج"'
            elif col.type == "percentage" and isinstance(val, (int, float)):
                cell.number_format = '0.00"%"'
            elif col.type == "number" and isinstance(val, (int, float)):
                cell.number_format = "#,##0.##"

            # Accumulate totals
            if col.type in ("number", "currency", "percentage", "formula") and isinstance(val, (int, float)):
                totals[col.key] = totals.get(col.key, 0.0) + val

        ws.row_dimensions[current_row].height = 20
        current_row += 1

    # ── Totals row ────────────────────────────────────────────────────────────
    if req.include_totals and req.rows:
        for ci, col in enumerate(req.columns, start=1):
            if ci == 1:
                cell = ws.cell(row=current_row, column=ci, value="الإجمالي")
            else:
                total_val = totals.get(col.key)
                cell = ws.cell(
                    row=current_row, column=ci,
                    value=total_val if col.type in ("number", "currency", "percentage", "formula") else ""
                )
                if col.type == "currency":
                    cell.number_format = '#,##0 "ج"'
                elif col.type == "percentage":
                    cell.number_format = '0.00"%"'
                elif col.type == "number":
                    cell.number_format = "#,##0.##"

            cell.font = TOTAL_FONT
            cell.fill = TOTAL_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                top=Side(style="double", color="2D5D89"),
                bottom=Side(style="thin", color="2D5D89"),
                left=thin_side,
                right=thin_side,
            )
        ws.row_dimensions[current_row].height = 24

    # ── Column widths ─────────────────────────────────────────────────────────
    for ci, col in enumerate(req.columns, start=1):
        # Auto-fit: measure max content length
        max_len = max(
            len(str(col.label)),
            *[len(str(row.cells.get(col.key, ""))) for row in req.rows],
            8
        )
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len * 1.4 + 4, 50)

    # ── Freeze header row ─────────────────────────────────────────────────────
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    # ── Stream response ───────────────────────────────────────────────────────
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{req.sheet_name}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
